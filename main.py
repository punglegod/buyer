import openpyxl
import os
import re
from copy import copy
import tkinter as tk
from tkinter import messagebox

# ================= 配置区域 =================
TEMPLATE_FILE = "抛单-段腾飞.xlsx"
SUPPLIER_MAPPING = {
    "台州乔克科技有限公司": "乔克",
    "顾家家居股份有限公司": "顾家",
    "浙江众望布艺股份有限公司": "众望",
    "众望布艺股份有限公司": "众望",
    "浙江玛雅布业有限公司": "玛雅",
    "嘉兴泰恩弹簧有限公司": "泰恩",
    "磨根国际贸易(上海)有限公司": "磨根",
    "浙江鑫飞腾塑料科技有限公司": "鑫飞腾",
    "天长市西达克家居有限公司": "西达克",
    "淄博邦世实业有限公司": "邦世",
    "江苏正川智能家居制造有限公司": "正川"
}

def get_supplier_short(full_name):
    if full_name in SUPPLIER_MAPPING:
        return SUPPLIER_MAPPING[full_name]
    short = re.sub(r'[（\(].*?[）\)]', '', full_name)
    
    # 移除常见的省份、城市等地名
    places = r'(广州|淄博|北京|上海|天津|重庆|河北|山西|辽宁|吉林|黑龙江|江苏|浙江|安徽|福建|江西|山东|河南|湖北|湖南|广东|海南|四川|贵州|云南|陕西|甘肃|青海|台湾|内蒙古|广西|西藏|宁夏|新疆|香港|澳门|广州|深圳|杭州|宁波|温州|嘉兴|湖州|绍兴|金华|衢州|舟山|台州|丽水|成都|东莞|佛山)'
    short = re.sub(places, '', short)
    
    # 移除常见的公司后缀和行业词
    short = re.sub(r'(纺织|家居|用品|金属|五金|智能|布艺|有限责任公司|有限公司|国际|实业|股份|科技|工业|贸易|中心|集团|布业|塑胶|家居|家具|包装|材料|电器|设备|LIMITED|LTD|CO\.|INC\.)',
                   '', short, flags=re.IGNORECASE)
    return short.strip()[:4]


# ================= 核心处理逻辑 =================
def process_order(input_text, log_func):
    orders_by_supplier = {}
    lines = input_text.strip().split('\n')

    # ====== 第一阶段：扫描并分组 ======
    for line in lines:
        raw_line = line.strip()
        if not raw_line: continue

        dept, order_id, date, supplier = "", "", "", ""
        item = {'mat_code': '', 'mat_desc': '', 'status': '', 'unit': '', 'qty': ''}

        # ====== 优先使用 Tab 分隔解析（更可靠） ======
        if '\t' in raw_line:
            fields = [f.strip() for f in raw_line.split('\t')]
            # 格式: 部门(0) 人名(1) 日期(2) 订单号(3) 物料代码(4) 物料描述(5) 状态(6) 单位(7) 数量(8) 交期(9) 供应商(10)
            if len(fields) >= 11:
                dept = fields[0]
                order_id = fields[3]
                item['mat_code'] = fields[4]
                item['mat_desc'] = fields[5]
                item['status'] = fields[6]
                item['unit'] = fields[7]
                item['qty'] = fields[8]
                date = fields[9]
                supplier = fields[10]

                if supplier and item['mat_code']:
                    if supplier not in orders_by_supplier:
                        orders_by_supplier[supplier] = {'dept': dept, 'order_id': order_id, 'date': date, 'items': []}
                    orders_by_supplier[supplier]['items'].append(item)
                continue

        # ====== 回退：空格分隔的正则解析 ======
        line = re.sub(r'\s+', ' ', raw_line).strip()

        dept_match = re.search(r'(固定产品研发中心|功能品类发展部)', line)
        if dept_match: dept = dept_match.group(1)

        order_match = re.search(r'(CP[A-Z0-9]{10,})', line)
        if order_match: order_id = order_match.group(1)

        dates = re.findall(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', line)
        if dates:
            date = dates[-1]
            last_idx = line.rfind(date)
            if last_idx != -1:
                supplier = line[last_idx + len(date):].strip()

        mat_code_match = re.search(r'(\d{2}\.\d{2}\.[A-Z0-9\-]+)', line)
        if mat_code_match: item['mat_code'] = mat_code_match.group(1)

        sqq_match = re.search(r'([A-Z0-9]{2,3})\s*([\u4e00-\u9fa5]+)\s*(\d+)', line)
        if sqq_match:
            item['status'] = sqq_match.group(1)
            item['unit'] = sqq_match.group(2)
            item['qty'] = sqq_match.group(3)

        if item['mat_code'] and item['status']:
            desc_pattern = re.escape(item['mat_code']) + r'\s*(.*?)\s*' + re.escape(item['status'])
            desc_match = re.search(desc_pattern, line)
            if desc_match: item['mat_desc'] = desc_match.group(1).strip()

        if supplier and item['mat_code']:
            if supplier not in orders_by_supplier:
                orders_by_supplier[supplier] = {'dept': dept, 'order_id': order_id, 'date': date, 'items': []}
            orders_by_supplier[supplier]['items'].append(item)

    if not orders_by_supplier:
        log_func("❌ 错误：未能识别到有效订单，请检查粘贴的格式是否正确。")
        return False

    # ====== 第二阶段：生成 Excel ======
    for supplier, order_data in orders_by_supplier.items():
        log_func(f"\n📦 开始处理供应商: {supplier} (共 {len(order_data['items'])} 条物料)")

        try:
            wb = openpyxl.load_workbook(TEMPLATE_FILE)
            ws = wb.active
        except FileNotFoundError:
            log_func(f"❌ 致命错误：找不到模板文件 '{TEMPLATE_FILE}'")
            log_func("👉 请确保模板文件与本程序放在同一个文件夹内！")
            return False

        r_gui, r_zhao = None, None
        for r in range(4, 15):
            cell_val = str(ws.cell(row=r, column=10).value or "")
            if "桂小强" in cell_val: r_gui = r
            if "赵必保" in cell_val: r_zhao = r

        dept_val = order_data['dept']
        if "固定" in dept_val and r_gui:
            if r_zhao:
                ws.cell(row=r_gui, column=10).value = ws.cell(row=r_zhao, column=10).value
                ws.cell(row=r_zhao, column=10).value = ""
            else:
                ws.cell(row=r_gui, column=10).value = ""
        elif "功能" in dept_val and r_zhao:
            ws.cell(row=r_zhao, column=10).value = ""

        for i, item in enumerate(order_data['items']):
            current_row = 5 + i
            ws.cell(row=current_row, column=1).value = order_data['order_id']
            ws.cell(row=current_row, column=2).value = item['mat_code']
            ws.cell(row=current_row, column=3).value = item['mat_desc']
            ws.cell(row=current_row, column=4).value = item['status']
            ws.cell(row=current_row, column=5).value = item['unit']
            ws.cell(row=current_row, column=6).value = int(item['qty']) if item['qty'].isdigit() else item['qty']
            ws.cell(row=current_row, column=7).value = order_data['date']
            ws.cell(row=current_row, column=8).value = supplier

            if current_row > 5:
                if ws.row_dimensions[5].height is not None:
                    ws.row_dimensions[current_row].height = ws.row_dimensions[5].height
                for col in range(1, 9):
                    source_cell = ws.cell(row=5, column=col)
                    target_cell = ws.cell(row=current_row, column=col)
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = copy(source_cell.number_format)
                        target_cell.alignment = copy(source_cell.alignment)

            log_func(f"   ✅ 写入成功: {item['mat_code']} ({item['qty']}{item['unit']})")

        sup_short = get_supplier_short(supplier)
        new_filename = f"{order_data['order_id']}-{sup_short}-段腾飞.xlsx"
        try:
            wb.save(new_filename)
            log_func(f"🎉 文件已生成: {new_filename}")
        except PermissionError:
            log_func(f"❌ 错误：文件 {new_filename} 正在被其他程序占用！")
            log_func("👉 请先在 Excel 中关闭它，然后重试。")
            return False

    return True


# ================= GUI 界面构建 (iOS 26 Liquid Glass) =================
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("采购订单助手")
        self.root.geometry("860x720")
        self.root.minsize(750, 650)
        self.root.configure(bg="#D2D0D6")

        # iOS 26 Liquid Glass 色系
        self.font_h1 = ("Microsoft YaHei", 22, "bold")
        self.font_h2 = ("Microsoft YaHei", 11)
        self.font_body = ("Microsoft YaHei", 10)
        self.font_small = ("Microsoft YaHei", 9)
        self.font_code = ("Consolas", 10)

        # 毛玻璃卡片色 (半透明白的视觉模拟)
        self.color_glass = "#E8EAF0"       # 冷灰调毛玻璃
        self.color_glass_input = "#F0F1F5" # 输入框底色
        self.color_text = "#FFFFFF"        # 亮文字 (渐变上)
        self.color_text_dark = "#1C1C1E"   # 深色文字 (卡片上)
        self.color_text_sub = "#8E8E93"    # 次要文字
        self.color_accent = "#007AFF"      # iOS 标志蓝
        self.color_accent_glow = "#409CFF" # 蓝光晕
        self.color_green = "#30D158"       # iOS 绿
        self.color_border = "#C7C7CC"      # 细边线

        self._last_input = ""  # 保存上次被清除的输入内容
        self.setup_ui()

        self.root.bind("<Configure>", self._on_configure)
        self.root.after(50, self._initial_draw)

    def _initial_draw(self):
        self.root.update_idletasks()
        self._draw_bg()

    def _on_configure(self, event):
        if event.widget == self.root:
            self.root.after_cancel(getattr(self, '_resize_after', ''))
            self._resize_after = self.root.after(30, self._draw_bg)

    # ---------- 多段渐变绘制 ----------
    def _draw_bg(self):
        w = self.root.winfo_width()
        h = self.root.winfo_height()
        if w <= 1 or h <= 1:
            return
        if getattr(self, '_lw', 0) == w and getattr(self, '_lh', 0) == h:
            return
        self._lw, self._lh = w, h

        self.canvas.delete("bg")

        # Apple 风格：柔和高级感渐变
        # 顶部: 暖银灰  中部: 淡薰衣紫  底部: 柔蓝银
        stops = [
            (0.0,  (210, 208, 214)),  # #D2D0D6 暖银灰
            (0.30, (196, 192, 208)),  # #C4C0D0 淡薰衣紫
            (0.60, (189, 212, 231)),  # #BDD4E7 柔天蓝
            (1.0,  (200, 213, 216)),  # #C8D5D8 银青
        ]

        step = 3
        for y in range(0, h, step):
            t = y / max(h, 1)
            # 找到当前 t 所在的区间
            for k in range(len(stops) - 1):
                if stops[k][0] <= t <= stops[k + 1][0]:
                    local_t = (t - stops[k][0]) / (stops[k + 1][0] - stops[k][0])
                    r = int(stops[k][1][0] + (stops[k + 1][1][0] - stops[k][1][0]) * local_t)
                    g = int(stops[k][1][1] + (stops[k + 1][1][1] - stops[k][1][1]) * local_t)
                    b = int(stops[k][1][2] + (stops[k + 1][1][2] - stops[k][1][2]) * local_t)
                    break
            color = f"#{r:02x}{g:02x}{b:02x}"
            self.canvas.create_rectangle(0, y, w, y + step, fill=color, outline=color, tags="bg")

        self.canvas.lower("bg")
        self.canvas.coords("card_win", w / 2, h / 2)
        cw = min(740, w - 60)
        ch = min(640, h - 60)
        self.canvas.itemconfig("card_win", width=cw, height=ch)

    # ---------- UI 搭建 ----------
    def setup_ui(self):
        self.canvas = tk.Canvas(self.root, highlightthickness=0, bg="#D2D0D6", bd=0)
        self.canvas.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        # 毛玻璃卡片 (用 Frame 模拟半透明效果)
        self.card = tk.Frame(self.canvas, bg=self.color_glass, bd=0,
                             highlightbackground="#FFFFFF", highlightthickness=1)
        self.canvas.create_window(430, 360, window=self.card, width=740, height=640, tags="card_win")

        # ===== 标题区 =====
        header = tk.Frame(self.card, bg=self.color_glass)
        header.pack(fill=tk.X, padx=36, pady=(28, 0))

        # 小标签 (模拟 iOS 小灰字标签)
        tk.Label(header, text="PROCUREMENT", font=("Microsoft YaHei", 9), 
                 bg=self.color_glass, fg=self.color_text_sub).pack(anchor=tk.W)

        tk.Label(header, text="采购订单助手", font=self.font_h1,
                 bg=self.color_glass, fg=self.color_text_dark).pack(anchor=tk.W, pady=(2, 0))

        tk.Label(header, text="智能拆单  ·  一键生成  ·  高效办公", font=self.font_h2,
                 bg=self.color_glass, fg=self.color_text_sub).pack(anchor=tk.W, pady=(3, 0))

        # 分割线
        sep = tk.Frame(self.card, bg=self.color_border, height=1)
        sep.pack(fill=tk.X, padx=36, pady=(18, 0))

        # ===== 输入区 =====
        body = tk.Frame(self.card, bg=self.color_glass, padx=36)
        body.pack(fill=tk.BOTH, expand=True, pady=(14, 0))

        tk.Label(body, text="订单数据", font=("Microsoft YaHei", 10, "bold"),
                 bg=self.color_glass, fg=self.color_text_dark).pack(anchor=tk.W, pady=(0, 5))

        # 输入框 — 浅色圆润感
        input_wrap = tk.Frame(body, bg=self.color_border, bd=0)
        input_wrap.pack(fill=tk.X)
        self.text_input = tk.Text(input_wrap, height=7, font=self.font_code, relief=tk.FLAT, bd=0,
                                  bg=self.color_glass_input, fg=self.color_text_dark,
                                  insertbackground=self.color_accent, selectbackground=self.color_accent_glow,
                                  padx=14, pady=10)
        self.text_input.pack(fill=tk.X, padx=1, pady=1)

        # ===== 按钮区 =====
        btn_frame = tk.Frame(body, bg=self.color_glass)
        btn_frame.pack(fill=tk.X, pady=(16, 0))

        # 主按钮 — iOS 蓝色胶囊
        self.run_btn = tk.Button(
            btn_frame, text="一键拆单生成", font=("Microsoft YaHei", 11, "bold"),
            bg=self.color_accent, fg="white",
            activebackground=self.color_accent_glow, activeforeground="white",
            relief=tk.FLAT, cursor="hand2", command=self.run_process, padx=28, pady=7)
        self.run_btn.pack(side=tk.LEFT)

        def _hover_in(e):
            if str(self.run_btn['state']) != 'disabled':
                self.run_btn.config(bg=self.color_accent_glow)
        def _hover_out(e):
            if str(self.run_btn['state']) != 'disabled':
                self.run_btn.config(bg=self.color_accent)
        self.run_btn.bind("<Enter>", _hover_in)
        self.run_btn.bind("<Leave>", _hover_out)

        # 辅助按钮 — 毛玻璃灰底
        def _make_glass_btn(parent, text, cmd):
            b = tk.Button(parent, text=text, font=self.font_body, relief=tk.FLAT, cursor="hand2",
                          bg="#DCDCE0", fg=self.color_text_dark, activebackground=self.color_border,
                          command=cmd, padx=14, pady=5)
            b.bind("<Enter>", lambda e: b.config(bg=self.color_border))
            b.bind("<Leave>", lambda e: b.config(bg="#DCDCE0"))
            return b

        _make_glass_btn(btn_frame, "找回输入", self.restore_input).pack(side=tk.LEFT, padx=12)
        _make_glass_btn(btn_frame, "清空输入",
                        lambda: self.text_input.delete(1.0, tk.END)).pack(side=tk.LEFT, padx=(0, 12))
        _make_glass_btn(btn_frame, "打开输出文件夹", self.open_folder).pack(side=tk.RIGHT)

        # 分割线
        sep2 = tk.Frame(body, bg=self.color_border, height=1)
        sep2.pack(fill=tk.X, pady=(16, 0))

        # ===== 日志区 =====
        tk.Label(body, text="运行日志", font=("Microsoft YaHei", 10, "bold"),
                 bg=self.color_glass, fg=self.color_text_dark).pack(anchor=tk.W, pady=(12, 5))

        log_wrap = tk.Frame(body, bg="#2C2C2E", bd=0)
        log_wrap.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        self.text_log = tk.Text(log_wrap, height=8, font=self.font_code,
                                bg="#2C2C2E", fg=self.color_green, relief=tk.FLAT, bd=0,
                                insertbackground=self.color_green, padx=14, pady=10)
        self.text_log.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
        self.text_log.config(state=tk.DISABLED)

    # ---------- 工具方法 ----------
    def log(self, message):
        self.text_log.config(state=tk.NORMAL)
        self.text_log.insert(tk.END, message + "\n")
        self.text_log.see(tk.END)
        self.text_log.config(state=tk.DISABLED)
        self.root.update()

    def restore_input(self):
        if self._last_input:
            self.text_input.delete(1.0, tk.END)
            self.text_input.insert(1.0, self._last_input)
            self.log("🔄 已找回上次输入内容。")
        else:
            self.log("⚠️ 没有可找回的历史输入。")

    def open_folder(self):
        try:
            os.startfile(os.getcwd())
        except Exception as e:
            self.log(f"⚠️ 无法自动打开文件夹: {e}")

    def run_process(self):
        content = self.text_input.get(1.0, tk.END).strip()
        if not content:
            messagebox.showwarning("提示", "您还没有粘贴任何订单内容哦！")
            return

        self.text_log.config(state=tk.NORMAL)
        self.text_log.delete(1.0, tk.END)
        self.text_log.config(state=tk.DISABLED)

        self.log("⏳ 开始解析数据...\n" + "-" * 40)
        self.run_btn.config(state=tk.DISABLED, text="处理中...", bg="#5A5A5E")
        self.root.update()

        success = process_order(content, self.log)

        self.log("-" * 40)
        if success:
            self.log("✨ 所有任务处理完毕！点击【打开输出文件夹】查看文件。")
            # 成功后自动清除输入，保存内容以便找回
            self._last_input = content
            self.text_input.delete(1.0, tk.END)
        else:
            self.log("⚠️ 处理过程中出现异常，请检查上述日志。")
            messagebox.showerror("失败", "处理过程中出现异常，请检查日志。")

        self.run_btn.config(state=tk.NORMAL, text="一键拆单生成", bg=self.color_accent)


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()